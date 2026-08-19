import openpyxl
import json

def read_excel_to_json(filepath):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                row_data[headers[i]] = cell.value
            data.append(row_data)
        return data
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return []

files = [
    '1级人物数据.xlsx',
    '2级人物数据.xlsx',
    '事件数据.xlsx',
    '人物关系表.xlsx',
    '人事关系表.xlsx'
]

all_data = {}
for file in files:
    data = read_excel_to_json(file)
    all_data[file.replace('.xlsx', '')] = data
    print(f"Read {len(data)} records from {file}")
    if data:
        print(f"Columns: {list(data[0].keys())}")

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print("Data saved to excel_data.json")